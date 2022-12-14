import openpyxl
import json
from openpyxl.cell.read_only import EmptyCell

workbook = openpyxl.load_workbook('hacknu-prod-data.xlsx')

cases = []
for sheet in workbook.worksheets:
    points = []
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row, column=1).value is None:
            break
        points.append({
            'Latitude':sheet.cell(row=row, column=1).value,
            'Longitude':sheet.cell(row=row, column=2).value,
            'Altitude':sheet.cell(row=row, column=3).value,
            'Identifier':sheet.cell(row=row, column=4).value,
            'Timestamp':sheet.cell(row=row, column=5).value,
            'Floor label':sheet.cell(row=row, column=6).value,
            'Horizontal accuracy':sheet.cell(row=row, column=7).value,
            'Vertical accuracy':sheet.cell(row=row, column=8).value,
            'Confidence in location accuracy':sheet.cell(row=row, column=9).value,
            'Activity':sheet.cell(row=row, column=10).value,
        })
    cases.append(points)
excel_json = open('assets/cases.json', 'w')
excel_json.write(json.dumps(cases))
excel_json.close()
